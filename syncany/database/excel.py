# -*- coding: utf-8 -*-
# 18/8/13
# create by: snower

import os
import datetime
import uuid
try:
    from bson.objectid import ObjectId
except ImportError:
    ObjectId = None
from ..utils import get_timezone, human_repr_object, sorted_by_keys
from ..taskers.context import TaskerContext
from ..taskers.iterator import TaskerDataIterator
from .database import Cmper, QueryBuilder, InsertBuilder, UpdateBuilder, DeleteBuilder, DataBase


class ExeclFileNotFound(Exception):
    pass


class ExeclQueryBuilder(QueryBuilder):
    def __init__(self, *args, **kwargs):
        super(ExeclQueryBuilder, self).__init__(*args, **kwargs)

    def filter_gt(self, key, value):
        self.query[(key, '>')] = (value, Cmper.cmp_gt)

    def filter_gte(self, key, value):
        self.query[(key, ">=")] = (value, Cmper.cmp_gte)

    def filter_lt(self, key, value):
        self.query[(key, "<")] = (value, Cmper.cmp_lt)

    def filter_lte(self, key, value):
        self.query[(key, "<=")] = (value, Cmper.cmp_lte)

    def filter_eq(self, key, value):
        self.query[(key, "==")] = (value, Cmper.cmp_eq)

    def filter_ne(self, key, value):
        self.query[(key, "!=")] = (value, Cmper.cmp_ne)

    def filter_in(self, key, value):
        try:
            self.query[(key, "in")] = (set(value) if isinstance(value, list) else value, Cmper.cmp_in)
        except:
            self.query[(key, "in")] = (value, Cmper.cmp_in)

    def filter_limit(self, count, start=None):
        if not start:
            self.limit = (0, count)
        else:
            self.limit = (start, start + count)

    def filter_cursor(self, last_data, offset, count, primary_orders=None):
        self.limit = (offset, offset + count)

    def order_by(self, key, direct=1):
        self.orders.append((key, direct))

    def commit(self):
        tasker_context, iterator_name, datas = None, None, None
        if self.limit and (self.query or self.orders):
            tasker_context = TaskerContext.current()
            if tasker_context:
                iterator_name = "excel::" + self.name
                iterator = tasker_context.get_iterator(iterator_name)
                if iterator and iterator.offset == self.limit[0]:
                    datas, iterator.offset = iterator.datas, self.limit[1]

        if not datas:
            execl_sheet = self.db.ensure_open_file(self.name)
            if not self.query:
                datas = execl_sheet.sheet_datas[:]
            else:
                datas = []
                for data in execl_sheet.sheet_datas:
                    succed = True
                    for (key, exp), (value, cmp) in self.query.items():
                        if key not in data:
                            succed = False
                            break
                        if not cmp(data[key], value):
                            succed = False
                            break
                    if succed:
                        datas.append(data)

            if self.orders:
                datas = sorted_by_keys(datas, keys=[(key, True if direct < 0 else False)
                                                    for key, direct in self.orders] if self.orders else None)
            if tasker_context and self.limit and (self.query or self.orders):
                tasker_context.add_iterator(iterator_name, TaskerDataIterator(datas, self.limit[1]))

        if self.limit:
            datas = datas[self.limit[0]: self.limit[1]]
        return datas

    def verbose(self):
        return "filters: %s\nlimit: %s\norderBy: %s" % (
            human_repr_object([(key, exp, value) for (key, exp), (value, cmp) in self.query.items()]),
            self.limit,
            self.orders)


class ExeclInsertBuilder(InsertBuilder):
    def __init__(self, *args, **kwargs):
        super(ExeclInsertBuilder, self).__init__(*args, **kwargs)

        if isinstance(self.datas, dict):
            self.datas = [self.datas]

    def commit(self):
        execl_sheet = self.db.ensure_open_file(self.name)
        execl_sheet.sheet_descriptions = self.fields
        execl_sheet.sheet_datas.extend(self.datas)
        execl_sheet.changed = True
        tasker_context = TaskerContext.current()
        if tasker_context:
            tasker_context.remove_iterator("excel::" + self.name)

    def verbose(self):
        return "datas(%d): \n%s" % (len(self.datas), human_repr_object(self.datas))


class ExeclUpdateBuilder(UpdateBuilder):
    def __init__(self, *args, **kwargs):
        super(ExeclUpdateBuilder, self).__init__(*args, **kwargs)

    def filter_gt(self, key, value):
        self.query[(key, '>')] = (value, Cmper.cmp_gt)

    def filter_gte(self, key, value):
        self.query[(key, ">=")] = (value, Cmper.cmp_gte)

    def filter_lt(self, key, value):
        self.query[(key, "<")] = (value, Cmper.cmp_lt)

    def filter_lte(self, key, value):
        self.query[(key, "<=")] = (value, Cmper.cmp_lte)

    def filter_eq(self, key, value):
        self.query[(key, "==")] = (value, Cmper.cmp_eq)

    def filter_ne(self, key, value):
        self.query[(key, "!=")] = (value, Cmper.cmp_ne)

    def filter_in(self, key, value):
        try:
            self.query[(key, "in")] = (set(value) if isinstance(value, list) else value, Cmper.cmp_in)
        except:
            self.query[(key, "in")] = (value, Cmper.cmp_in)

    def commit(self):
        execl_sheet = self.db.ensure_open_file(self.name)
        execl_sheet.sheet_descriptions = self.fields
        datas = []
        for data in execl_sheet.sheet_datas:
            succed = True
            for (key, exp), (value, cmp) in self.query.items():
                if key not in data:
                    succed = False
                    break
                if not cmp(data[key], value):
                    succed = False
                    break

            if succed:
                datas.append(self.update)
            else:
                datas.append(data)

        execl_sheet.sheet_datas = datas
        execl_sheet.changed = True
        tasker_context = TaskerContext.current()
        if tasker_context:
            tasker_context.remove_iterator("excel::" + self.name)
        return datas

    def verbose(self):
        return "filters: %s\nupdateDatas: %s" % (
            human_repr_object([(key, exp, value) for (key, exp), (value, cmp) in self.query.items()]),
            human_repr_object(self.diff_data))


class ExeclDeleteBuilder(DeleteBuilder):
    def __init__(self, *args, **kwargs):
        super(ExeclDeleteBuilder, self).__init__(*args, **kwargs)

    def filter_gt(self, key, value):
        self.query[(key, '>')] = (value, Cmper.cmp_gt)

    def filter_gte(self, key, value):
        self.query[(key, ">=")] = (value, Cmper.cmp_gte)

    def filter_lt(self, key, value):
        self.query[(key, "<")] = (value, Cmper.cmp_lt)

    def filter_lte(self, key, value):
        self.query[(key, "<=")] = (value, Cmper.cmp_lte)

    def filter_eq(self, key, value):
        self.query[(key, "==")] = (value, Cmper.cmp_eq)

    def filter_ne(self, key, value):
        self.query[(key, "!=")] = (value, Cmper.cmp_ne)

    def filter_in(self, key, value):
        try:
            self.query[(key, "in")] = (set(value) if isinstance(value, list) else value, Cmper.cmp_in)
        except:
            self.query[(key, "in")] = (value, Cmper.cmp_in)

    def commit(self):
        execl_sheet = self.db.ensure_open_file(self.name)
        datas = []
        for data in execl_sheet.sheet_datas:
            succed = True
            for (key, exp), (value, cmp) in self.query.items():
                if key not in data:
                    succed = False
                    break
                if not cmp(data[key], value):
                    succed = False
                    break

            if not succed:
                datas.append(data)

        execl_sheet.sheet_datas = datas
        execl_sheet.changed = True
        tasker_context = TaskerContext.current()
        if tasker_context:
            tasker_context.remove_iterator("excel::" + self.name)
        return datas

    def verbose(self):
        return "filters: %s" % human_repr_object([(key, exp, value) for (key, exp), (value, cmp) in self.query.items()])


class ExeclSheet(object):
    def __init__(self, name, filename, sheet_name, execl_fp, execl_sheet):
        self.name = name
        self.filename = filename
        self.sheet_name = sheet_name

        self.execl_fp = execl_fp
        self.execl_sheet = execl_sheet
        self.sheet_descriptions = []
        self.sheet_datas = []
        self.changed = False

    def load(self):
        for row in self.execl_sheet.rows:
            if not self.sheet_descriptions:
                for cel in row:
                    self.sheet_descriptions.append(cel.value)
            else:
                data, index = {}, 0
                for cel in row:
                    if isinstance(cel.value, datetime.datetime) and not cel.value.tzinfo:
                        data[self.sheet_descriptions[index]] = cel.value.replace(tzinfo=get_timezone())
                    else:
                        data[self.sheet_descriptions[index]] = cel.value
                    index += 1
                self.sheet_datas.append(data)

    def get_fields(self):
        if self.sheet_descriptions:
            return self.sheet_descriptions

        fields = None
        for data in self.sheet_datas:
            if fields is None:
                fields = set(data.keys())
            else:
                fields = fields & set(data.keys())
        return tuple(fields) if fields else tuple()

    def close(self):
        if not self.changed:
            return

        if self.execl_sheet:
            sheet_index = self.execl_fp.index(self.execl_sheet)
            self.execl_fp.remove(self.execl_sheet)
            self.execl_sheet = self.execl_fp.create_sheet(self.sheet_name, sheet_index)
            fields = self.get_fields()

            for field_index in range(len(fields)):
                field = fields[field_index]
                if isinstance(field, str):
                    field = field.encode("utf-8")
                self.execl_sheet.cell(1, field_index + 1, field)

            for row_index in range(len(self.sheet_datas)):
                row = self.sheet_datas[row_index]
                for col_index in range(len(fields)):
                    col = row[fields[col_index]]
                    if isinstance(col, datetime.datetime):
                        if col.tzinfo:
                            col = col.replace(tzinfo=None)
                    elif ObjectId is not None and isinstance(col, ObjectId):
                        col = str(col)
                    elif isinstance(col, uuid.UUID):
                        col = str(col)
                    if isinstance(col, str):
                        col = col.encode("utf-8")
                    self.execl_sheet.cell(row_index + 2, col_index + 1, col)


class ExeclDB(DataBase):
    DEFAULT_CONFIG = {
        "path": "./",
    }

    def __init__(self, manager, config):
        all_config = {}
        all_config.update(self.DEFAULT_CONFIG)
        all_config.update(config)

        all_config["path"] = os.path.abspath(all_config["path"])

        super(ExeclDB, self).__init__(manager, all_config)

        self.execls = {}
        self.execl_fps = {}

    def parse_filename(self, name):
        filename, sheet_name = None, None
        names = name.split(".")
        if len(names) >= 2:
            filenames = ".".join(names[1:]).split("#")
            if len(filenames) >= 2:
                filename = filenames[0]
                sheet_name = filenames[1]
            else:
                filename = filenames[0]
        return filename, sheet_name

    def ensure_open_file(self, name):
        if not name:
            raise ExeclFileNotFound()

        if name not in self.execls:
            filename, sheet_name = self.parse_filename(name)
            if not filename:
                raise ExeclFileNotFound()

            try:
                import openpyxl
            except ImportError:
                raise ImportError("openpyxl>=2.5.0 is required")

            filename = os.path.join(self.config["path"], filename)
            if filename in self.execl_fps:
                execl_fp = self.execl_fps[filename]
            elif os.path.exists(filename):
                self.execl_fps[filename] = execl_fp = openpyxl.load_workbook(filename)
            else:
                execl_fp = None

            if execl_fp:
                if sheet_name:
                    execl_sheet = None
                    for sheet in execl_fp.worksheets:
                        if sheet.title == sheet_name:
                            execl_sheet = sheet
                            break

                    if not execl_sheet:
                        execl_sheet = execl_fp.create_sheet(sheet_name)
                else:
                    worksheets = execl_fp.worksheets
                    if not worksheets:
                        execl_sheet = execl_fp.create_sheet()
                    else:
                        execl_sheet = worksheets[0]
            else:
                self.execl_fps[filename] = execl_fp = openpyxl.Workbook()
                execl_sheet = execl_fp.worksheets[0]
                if sheet_name:
                    execl_sheet.title = sheet_name

            self.execls[name] = ExeclSheet(name, filename, sheet_name, execl_fp, execl_sheet)
            self.execls[name].load()

        return self.execls[name]

    def query(self, name, primary_keys=None, fields=()):
        return ExeclQueryBuilder(self, name, primary_keys, fields)

    def insert(self, name, primary_keys=None, fields=(), datas=None):
        return ExeclInsertBuilder(self, name, primary_keys, fields, datas)

    def update(self, name, primary_keys=None, fields=(), update=None, diff_data=None):
        return ExeclUpdateBuilder(self, name, primary_keys, fields, update, diff_data)

    def delete(self, name, primary_keys=None):
        return ExeclDeleteBuilder(self, name, primary_keys)

    def flush(self):
        if self.execls:
            for name, execl_sheet in self.execls.items():
                execl_sheet.close()

        if self.execl_fps:
            for filename, execl_fp in self.execl_fps.items():
                execl_fp.save(filename)

        self.execls = {}
        self.execl_fps = {}

    def close(self):
        self.flush()

    def verbose(self):
        return "%s<%s>" % (self.name, self.config["path"])